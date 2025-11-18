# File: backtester.py
# Description: Complete backtester with multiprocessing, comprehensive reports, and RL environment consistency
# =============================================================================
import pandas as pd
import numpy as np
import logging
import os
import glob
import json
from openpyxl import load_workbook
from tqdm import tqdm
from stable_baselines3 import PPO
import multiprocessing
from functools import partial
from datetime import datetime
import matplotlib.pyplot as plt
import traceback

import config
import reporting
from rl_environment import TradingEnv

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def safe_value(x, default=0.0, clip_min=-1e12, clip_max=1e12):
    """Sanitize values: replace NaN/Inf with default, and clip extreme values."""
    if np.isnan(x) or np.isinf(x):
        return default
    return float(np.clip(x, clip_min, clip_max))


def run_rl_backtest_with_tracking(data_df, model, ticker, model_name=""):
    """
    Run backtest using TradingEnv with detailed trade tracking.
    
    This ensures:
    - 100% consistency with training
    - Realistic timing (observe completed candles, execute on next)
    - Detailed trade capture via monkey-patching
    - Comprehensive equity tracking
    """
    try:
        # Create environment
        env = TradingEnv(data_df)
        
        # Setup trade tracking via monkey-patching
        detailed_trades = []
        original_update = env._update_portfolio
        
        def tracked_update():
            """Intercept position updates to capture trade details"""
            positions_before = [p.copy() for p in env.open_positions]
            pnl = original_update()
            
            # Find closed positions
            positions_after_ids = [id(p) for p in env.open_positions]
            closed_positions = [p for p in positions_before if id(p) not in positions_after_ids]
            
            if closed_positions and pnl != 0:
                current_idx = max(0, min(len(env.close_prices) - 1, int(env.current_step - 1)))
                candle_high = env.df['High'].iloc[current_idx]
                candle_low  = env.df['Low'].iloc[current_idx]
                current_time = data_df.index[current_idx]

                for pos in closed_positions:
                    entry_price = pos['price']
                    sl = pos['sl']
                    tp = pos['tp']
                    size = pos['size']
                    pos_type = pos['type']

                    if pos_type == 'LONG':
                        # If candle touches both SL and TP, SL first
                        if candle_low <= sl and candle_high >= tp:
                            exit_price = sl
                            exit_reason = 'SL'
                        elif candle_low <= sl:
                            exit_price = sl
                            exit_reason = 'SL'
                        elif candle_high >= tp:
                            exit_price = tp
                            exit_reason = 'TP'
                        else:
                            continue  # Should not happen

                        trade_pnl = (exit_price - entry_price) * size

                    else:  # SHORT
                        # If both levels touched, SL first (worst)
                        if candle_high >= sl and candle_low <= tp:
                            exit_price = sl
                            exit_reason = 'SL'
                        elif candle_high >= sl:
                            exit_price = sl
                            exit_reason = 'SL'
                        elif candle_low <= tp:
                            exit_price = tp
                            exit_reason = 'TP'
                        else:
                            continue  # Should not happen

                        trade_pnl = (entry_price - exit_price) * size

                    trade_pnl = safe_value(trade_pnl)
                    position_value = entry_price * size
                    pnl_percent = (trade_pnl / env.equity) if position_value > 0 else 0

                    detailed_trades.append({
                        'type': pos_type,
                        'price': entry_price,
                        'close_price': exit_price,
                        'close_time': current_time,
                        'sl': sl,
                        'tp': tp,
                        'size': size,
                        'pnl': trade_pnl,
                        'pnl_percent': pnl_percent,
                        'exit_reason': exit_reason,
                        'risk_level': pos.get('risk_level'),
                        'rr_profile_index': pos.get('rr_profile_index')
                    })
            return pnl
        
        env._update_portfolio = tracked_update
        
        # Reset and run backtest
        obs, _ = env.reset()
        equity_curve = []
        total_steps = len(data_df) - env.lookback_window - 1
        
        for step in range(total_steps):
            try:
                # Record equity
                current_time = data_df.index[env.current_step - 1] if env.current_step > 0 else None
                equity_curve.append({
                    'Time': current_time,
                    'Equity': safe_value(env.equity)
                })
                
                # Get action and step
                action, _ = model.predict(obs, deterministic=True)
                obs, reward, terminated, truncated, _ = env.step(action)
                
                # Sanitize observation
                if isinstance(obs, dict):
                    for key in obs:
                        obs[key] = np.nan_to_num(obs[key], nan=0.0, posinf=1e9, neginf=-1e9)
                else:
                    obs = np.nan_to_num(obs, nan=0.0, posinf=1e9, neginf=-1e9)
                
                if terminated or truncated:
                    break
                    
            except Exception as e:
                logger.error(f"Error at step {step} for {model_name}: {e}")
                break
        
        # Add final equity point
        if env.current_step > 0 and env.current_step <= len(data_df):
            final_time = data_df.index[env.current_step - 1]
            equity_curve.append({
                'Time': final_time,
                'Equity': safe_value(env.equity)
            })
        
        # Calculate max drawdown
        equity_df = pd.DataFrame(equity_curve)
        if not equity_df.empty and equity_df['Equity'].max() > 0:
            peak = equity_df['Equity'].cummax()
            drawdown = (equity_df['Equity'] - peak) / peak
            max_drawdown = -safe_value(drawdown.min(), default=0) * 100
        else:
            max_drawdown = 0
        
        trades_df = pd.DataFrame(detailed_trades)
        
        return equity_df, trades_df, max_drawdown
        
    except Exception as e:
        logger.error(f"Fatal error in backtest for {model_name}: {e}")
        logger.error(traceback.format_exc())
        return pd.DataFrame(), pd.DataFrame(), 0


def generate_trade_breakdown(trades_df, model_name):
    """Generate detailed breakdown of trades by risk level and RR profile"""
    breakdown = {
        'model_name': model_name,
        'total_trades': len(trades_df),
        'by_risk_level': {},
        'by_rr_profile': {}
    }
    
    if trades_df.empty:
        return breakdown
    
    # Breakdown by Risk Level
    if 'risk_level' in trades_df.columns:
        risk_groups = trades_df.groupby('risk_level')
        for risk_level, group in risk_groups:
            wins = (group['pnl'] > 0).sum()
            total_trades = len(group)
            win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
            total_pnl = group['pnl'].sum()
            avg_pnl = group['pnl'].mean()
            
            breakdown['by_risk_level'][str(risk_level)] = {
                'trades': total_trades,
                'wins': wins,
                'losses': total_trades - wins,
                'win_rate_pct': round(win_rate, 2),
                'total_pnl': round(total_pnl, 2),
                'avg_pnl': round(avg_pnl, 2)
            }
    
    # Breakdown by RR Profile
    if 'rr_profile_index' in trades_df.columns:
        rr_groups = trades_df.groupby('rr_profile_index')
        for rr_index, group in rr_groups:
            wins = (group['pnl'] > 0).sum()
            total_trades = len(group)
            win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
            total_pnl = group['pnl'].sum()
            avg_pnl = group['pnl'].mean()
            
            # Get RR profile details
            rr_index_int = int(rr_index)
            if rr_index_int < len(config.RR_PROFILES):
                sl_mult, tp_mult = config.RR_PROFILES[rr_index_int]
                rr_ratio = f"1:{tp_mult/sl_mult:.1f}" if sl_mult != 0 else "N/A"
            else:
                sl_mult, tp_mult = 0, 0
                rr_ratio = "Unknown"
            
            breakdown['by_rr_profile'][str(rr_index)] = {
                'sl_multiplier': sl_mult,
                'tp_multiplier': tp_mult,
                'rr_ratio': rr_ratio,
                'trades': total_trades,
                'wins': wins,
                'losses': total_trades - wins,
                'win_rate_pct': round(win_rate, 2),
                'total_pnl': round(total_pnl, 2),
                'avg_pnl': round(avg_pnl, 2)
            }
    
    return breakdown


def save_comprehensive_results(all_results, ticker):
    """Save comprehensive results for all models in a single Excel file"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    results_dir = os.path.join("backtest_results")
    os.makedirs(results_dir, exist_ok=True)
    
    excel_file = os.path.join(results_dir, f"{ticker}_backtest_results_{timestamp}.xlsx")
    
    # Prepare data for Excel
    summary_data = []
    all_trades_data = []
    all_equity_data = []
    risk_breakdown_data = []
    rr_breakdown_data = []
    
    for res in all_results:
        model_name = res['model_name']
        
        # Summary data
        summary_data.append({
            'Model': model_name,
            **res['metrics']
        })
        
        # All trades with model identifier
        if not res['trades'].empty:
            trades_with_model = res['trades'].copy()
            trades_with_model['Model'] = model_name
            all_trades_data.append(trades_with_model)
            
            # Generate trade breakdowns
            breakdown = generate_trade_breakdown(res['trades'], model_name)
            
            # Risk level breakdown
            for risk_level, data in breakdown['by_risk_level'].items():
                risk_breakdown_data.append({
                    'Model': model_name,
                    'Risk_Level_%': risk_level,
                    'Trades': data['trades'],
                    'Wins': data['wins'],
                    'Losses': data['losses'],
                    'Win_Rate_%': data['win_rate_pct'],
                    'Total_PnL': data['total_pnl'],
                    'Avg_PnL': data['avg_pnl']
                })
            
            # RR profile breakdown
            for rr_index, data in breakdown['by_rr_profile'].items():
                rr_breakdown_data.append({
                    'Model': model_name,
                    'RR_Profile_Index': rr_index,
                    'SL_Multiplier': data['sl_multiplier'],
                    'TP_Multiplier': data['tp_multiplier'],
                    'RR_Ratio': data['rr_ratio'],
                    'Trades': data['trades'],
                    'Wins': data['wins'],
                    'Losses': data['losses'],
                    'Win_Rate_%': data['win_rate_pct'],
                    'Total_PnL': data['total_pnl'],
                    'Avg_PnL': data['avg_pnl']
                })
        
        # Equity curves with model identifier
        if not res['equity_curve'].empty:
            equity_with_model = res['equity_curve'].copy()
            equity_with_model['Model'] = model_name
            all_equity_data.append(equity_with_model)
    
    # Create comprehensive Excel file
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 1. Summary sheet - sorted by Calmar Ratio
            summary_df = pd.DataFrame(summary_data)
            summary_df = summary_df.sort_values(by="Calmar Ratio", ascending=False)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # 2. All trades in one sheet
            if all_trades_data:
                all_trades_df = pd.concat(all_trades_data, ignore_index=True)
                cols = ['Model'] + [col for col in all_trades_df.columns if col != 'Model']
                all_trades_df = all_trades_df[cols]
                all_trades_df.to_excel(writer, sheet_name='All_Trades', index=False)
            
            # 3. All equity curves in one sheet
            if all_equity_data:
                all_equity_df = pd.concat(all_equity_data, ignore_index=True)
                cols = ['Model'] + [col for col in all_equity_df.columns if col != 'Model']
                all_equity_df = all_equity_df[cols]
                all_equity_df.to_excel(writer, sheet_name='All_Equity_Curves', index=False)
            
            # 4. Risk level breakdown for all models
            if risk_breakdown_data:
                risk_df = pd.DataFrame(risk_breakdown_data)
                risk_df.to_excel(writer, sheet_name='Risk_Level_Breakdown', index=False)
            
            # 5. RR profile breakdown for all models
            if rr_breakdown_data:
                rr_df = pd.DataFrame(rr_breakdown_data)
                rr_df.to_excel(writer, sheet_name='RR_Profile_Breakdown', index=False)
            
            # 6. Metadata sheet
            metadata_data = {
                'Parameter': [
                    'Timestamp', 'Ticker', 'Total Models Tested', 'Initial Equity',
                    'Risk Levels Used', 'RR Profiles Used', 'Lookback Window'
                ],
                'Value': [
                    timestamp, ticker, len(all_results), config.INITIAL_EQUITY,
                    str(config.RISK_LEVELS), str(config.RR_PROFILES), config.RL_LOOKBACK_WINDOW
                ]
            }
            
            if all_results and not all_results[0]['equity_curve'].empty:
                metadata_data['Parameter'].extend(['Backtest Start', 'Backtest End'])
                metadata_data['Value'].extend([
                    str(min([res['equity_curve']['Time'].min() for res in all_results if not res['equity_curve'].empty])),
                    str(max([res['equity_curve']['Time'].max() for res in all_results if not res['equity_curve'].empty]))
                ])
            
            metadata_df = pd.DataFrame(metadata_data)
            metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
            
            # 7. Top performers summary
            if len(summary_df) > 0:
                top_performers_data = []
                metrics_to_rank = ['Calmar Ratio', 'Profit Factor', 'Total Return (%)', 'Win Rate (%)', 'Sharpe Ratio (Annualized)']
                
                for metric in metrics_to_rank:
                    if metric in summary_df.columns:
                        top_10 = summary_df.nlargest(min(10, len(summary_df)), metric)[['Model', metric]].copy()
                        top_10['Rank'] = range(1, len(top_10) + 1)
                        top_10['Metric'] = metric
                        top_10 = top_10[['Metric', 'Rank', 'Model', metric]]
                        top_performers_data.append(top_10)
                
                if top_performers_data:
                    top_performers_df = pd.concat(top_performers_data, ignore_index=True)
                    top_performers_df.to_excel(writer, sheet_name='Top_Performers', index=False)
        
        logger.info(f"✅ Comprehensive Excel file saved: {excel_file}")
        
        # Auto-adjust column widths
        try:
            wb = load_workbook(excel_file)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_file)
            logger.info("✅ Excel column widths auto-adjusted")
            
        except Exception as e:
            logger.warning(f"Could not auto-adjust Excel column widths: {e}")
        
    except Exception as e:
        logger.error(f"Could not create Excel report: {e}")
        logger.error(traceback.format_exc())
        return None
    
    return excel_file


def backtest_worker(model_path, data_path, ticker):
    """Worker function for multiprocessing - loads model and runs backtest"""
    try:
        model_name = os.path.basename(model_path)
        logger.info(f"Starting backtest for: {model_name}")
        
        # Load data
        backtest_df = pd.read_parquet(data_path)
        
        # Load model
        model = PPO.load(model_path, device="cpu")
        
        # Run backtest with tracking
        equity_curve, trades, max_drawdown = run_rl_backtest_with_tracking(
            backtest_df, model, ticker, model_name
        )
        
        if not equity_curve.empty:
            metrics = reporting.calculate_metrics(equity_curve, trades, max_drawdown)
            logger.info(f"✅ Completed: {model_name} | Calmar: {metrics.get('Calmar Ratio', 0):.2f}")
            return {
                "model_name": model_name,
                "metrics": metrics,
                "equity_curve": equity_curve,
                "trades": trades
            }
        else:
            logger.warning(f"⚠️  No results for: {model_name}")
            return None
            
    except Exception as e:
        logger.error(f"❌ Error backtesting {os.path.basename(model_path)}: {e}")
        logger.error(traceback.format_exc())
        return None

def save_best_model_name(model_name: str, file_path: str = "best_model.json"):
    """
    Save the best model name to a JSON file.
    
    Args:
        model_name (str): The name or path of the best model.
        file_path (str): File path for the JSON file (default: 'best_model.json').
    """
    data = {"best_model": model_name}
    
    try:
        with open(file_path, "w") as f:
            json.dump(data, f, indent=4)
        print(f"✅ Best model name saved to '{file_path}': {model_name}")
    except Exception as e:
        print(f"❌ Failed to save best model name: {e}")

def main():
    """Main backtesting function with multiprocessing"""
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 1000)
    
    logger.info("=" * 80)
    logger.info("🚀 COMPREHENSIVE RL BACKTESTER")
    logger.info("=" * 80)
    
    ticker = config.TICKERS[0]
    
    # Check for processed data
    processed_data_path = os.path.join("processed_data", f"{ticker}_processed.parquet")
    if not os.path.exists(processed_data_path):
        logger.critical(f"❌ Processed data not found at {processed_data_path}")
        logger.critical("Please run preprocess_data.py first.")
        return
    
    logger.info(f"✅ Found processed data: {processed_data_path}")
    
    # Find all model files
    model_dir = config.RL_MODEL_DIR
    model_paths = [
        p for p in glob.glob(os.path.join(model_dir, "*.zip"))
    ]
    
    if not model_paths:
        logger.error(f"❌ No candidate models found in '{model_dir}'")
        logger.info("Looking for files matching: *.zip (excluding 'final')")
        return
    
    logger.info(f"✅ Found {len(model_paths)} candidate models")
    
    # Setup multiprocessing
    num_processes = min(8, os.cpu_count())
    logger.info(f"🔄 Using {num_processes} parallel processes")
    logger.info("=" * 80)
    
    # Run backtests in parallel
    try:
        with multiprocessing.Pool(processes=num_processes) as pool:
            worker_func = partial(backtest_worker, data_path=processed_data_path, ticker=ticker)
            all_results = [
                result for result in tqdm(
                    pool.imap_unordered(worker_func, model_paths),
                    total=len(model_paths),
                    desc="📊 Backtesting Models",
                    ncols=100
                ) if result is not None
            ]
    except Exception as e:
        logger.error(f"❌ Error during parallel backtesting: {e}")
        logger.error(traceback.format_exc())
        return
    
    if not all_results:
        logger.error("❌ No models produced valid backtest results")
        return
    
    logger.info(f"\n✅ Successfully backtested {len(all_results)}/{len(model_paths)} models")
    
    # Sort results by Calmar Ratio
    all_results.sort(key=lambda x: x['metrics'].get("Calmar Ratio", -999), reverse=True)
    
    # Save comprehensive results
    excel_path = save_comprehensive_results(all_results, ticker)
    
    # Display summary
    summary_data = []
    for res in all_results:
        summary_data.append({'model_name': res['model_name'], **res['metrics']})
    
    results_df = pd.DataFrame(summary_data)
    
    print("\n" + "="*120)
    print("📊 FULL BACKTEST SUMMARY (Top 10)")
    print("="*120)
    summary_columns = [
        'model_name', 'Calmar Ratio', 'Profit Factor', 'Total Return (%)',
        'Winning Months (%)', 'Win Rate (%)', 'Total Trades', 'Max Drawdown (%)'
    ]
    print(results_df.head(10).reindex(columns=summary_columns).fillna(0).to_string(index=False))
    print("="*120)
    
    # Best model details
    if not results_df.empty:
        best_model_info = results_df.iloc[0]
        best_result = all_results[0]
        
        save_best_model_name(best_model_info['model_name'])
        print(f"\n🏆 BEST MODEL: {best_model_info['model_name']}")
        print("=" * 80)
        print(f"  📈 Calmar Ratio: {best_model_info['Calmar Ratio']:.2f}")
        print(f"  💰 Profit Factor: {best_model_info['Profit Factor']:.2f}")
        print(f"  📊 Total Return: {best_model_info['Total Return (%)']:.2f}%")
        print(f"  ✅ Win Rate: {best_model_info['Win Rate (%)']:.2f}%")
        print(f"  📉 Max Drawdown: {best_model_info['Max Drawdown (%)']:.2f}%")
        print(f"  🎯 Total Trades: {int(best_model_info['Total Trades'])}")
        print("=" * 80)
        
        # Show breakdown for best model
        if not best_result['trades'].empty:
            breakdown_report = reporting.format_performance_breakdown(best_result['trades'])
            print(breakdown_report)
        
        print(f"\n💡 ACTION: Consider using '{best_model_info['model_name']}' for live trading")
        #print(f"📁 Results saved to: {excel_path}")
        #print(f"📊 Excel includes: Summary, All Trades, Equity Curves, Risk/RR Breakdowns, Top Performers")
        
        # Plot equity curve for best model
        """try:     
            equity_curve = best_result['equity_curve']
            plt.figure(figsize=(14, 7))
            plt.plot(equity_curve['Time'], equity_curve['Equity'], label='Equity Curve', linewidth=2, color='blue')
            plt.axhline(y=config.INITIAL_EQUITY, color='red', linestyle='--', alpha=0.5, label='Initial Equity')
            plt.title(f"Equity Curve - Best Model: {best_model_info['model_name']}", fontsize=14, fontweight='bold')
            plt.xlabel("Time", fontsize=12)
            plt.ylabel("Equity ($)", fontsize=12)
            plt.legend(fontsize=10)
            plt.grid(True, alpha=0.3)
            plt.tight_layout()
            
            plot_path = os.path.join("backtest_results", f"best_model_equity_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
            plt.savefig(plot_path, dpi=150)
            logger.info(f"📈 Best model equity plot saved: {plot_path}")
            plt.show()
        except Exception as e:
            logger.warning(f"Could not generate plot: {e}")"""
    
    logger.info("\n" + "=" * 80)
    logger.info("✅ BACKTESTING COMPLETE!")
    logger.info("=" * 80)


if __name__ == '__main__':
    # Setup multiprocessing
    try:
        multiprocessing.set_start_method('spawn', force=True)
    except RuntimeError:
        pass
    
    main()